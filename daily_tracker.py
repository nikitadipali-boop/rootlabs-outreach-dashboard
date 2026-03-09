"""
Daily Outreach Tracker
======================
Run this script once per day to:
  1. Pull the latest state from Airtable
  2. Compare to yesterday's snapshot to detect changes
  3. Append new entries to the changelog CSV
  4. Regenerate the daily summary Excel

Usage:
    python3 daily_tracker.py

Output files (all in the same folder as this script):
    snapshots/snapshot_YYYY-MM-DD.json   -- daily state snapshot
    changelog.csv                         -- every state transition ever detected
    daily_summary.xlsx                    -- Excel dashboard with inbox + execution metrics
"""

import json
import os
import csv
import time
import requests
from datetime import datetime, timezone, timedelta
from collections import defaultdict, Counter

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system(f"pip install openpyxl -q")
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────────────────────

TOKEN    = "pat0aSErPoCgOSR2B.4bde5ea5bcf124ac0680d144183be4baf5d158be0d19777e8a4fc7dd43037fa8"
BASE_ID  = "appnhGIoeLSfLf9ah"
TABLE_ID = "tblwZwNeuZwtIavqj"
HEADERS  = {"Authorization": f"Bearer {TOKEN}"}

BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
SNAPSHOT_DIR    = os.path.join(BASE_DIR, "snapshots")
CHANGELOG_PATH  = os.path.join(BASE_DIR, "changelog.csv")
SUMMARY_PATH    = os.path.join(BASE_DIR, "daily_summary.xlsx")
SCORECARD_PATH  = os.path.join(BASE_DIR, "daily_scorecards.csv")

os.makedirs(SNAPSHOT_DIR, exist_ok=True)

SCORECARD_FIELDS = [
    "date", "saved_at",
    "sod_needs_reply", "sod_needs_followup_1", "sod_needs_followup_2", "sod_needs_followup_3", "sod_total",
    "new_inbounds",
    "replied", "followup_1_sent", "followup_2_sent", "followup_3_sent", "total_actioned",
    "reply_rate_pct", "fu1_rate_pct", "fu2_rate_pct", "fu3_rate_pct", "overall_rate_pct",
]

NON_CREATOR_DOMAINS = {
    "periskope.app","accounts.google.com","apollo.io","mail.apollo.io","mail.anthropic.com",
    "tello.com","github.com","airtable.com","mail.airtable.com","engage.canva.com",
    "twelvelabs.io","notify.railway.app","news.railway.app","supabase.com","gamma.app",
    "apify.com","email.openai.com","tm.openai.com","email.claude.com","mail.respond.io",
    "reply.io","fyxer.com","mermaidchart.com","mermaid.ai","team.twilio.com","vidyard.com",
    "superagent.com","info.n8n.io","useloom.com","amazon.com","amazonaws.com",
    "lemlist-news.com","notifications.hubspot.com","mailchimp.com","send.zapier.com",
    "discord.com","reacherapp.com","klaviyo.com","boxbe.com","google.com",
    "successgncapital.com","qualfon.com","goshipcentrlpro.com","ibramdawwa-gmbh.com",
    "partnerssalesbytomorrowlead.info","huntdmfirm.info","evolvedcommerceflows.info",
    "meetapprovalprocessesdigital.com","geniusecommerce-today.com","checkaxisbrands.org",
}

INBOXES = [
    "may_k@rootlabs.co", "may.k@rootlabs.co", "founder@rootlabs.co",
    "may.kumar@rootlabs.co", "mayank.k@rootlabs.co", "mayank.kumar@rootlabs.co",
    "may@rootlabs.co", "ceo@rootlabs.co", "mayk@rootlabs.co",
]

# ── Styles ────────────────────────────────────────────────────────────────────

HEADER_FONT   = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL   = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ALT_FILL      = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
GREEN_FILL    = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
AMBER_FILL    = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL      = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
TOTAL_FILL    = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
DATA_FONT     = Font(name="Arial", size=10)
BOLD_FONT     = Font(name="Arial", size=10, bold=True)
THIN_BORDER   = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)


def style_sheet(ws, headers, data_rows, row_fills=None):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    for r_idx, row_data in enumerate(data_rows, 2):
        fill = row_fills[r_idx - 2] if row_fills else (ALT_FILL if r_idx % 2 == 0 else None)
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill

    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for row in range(2, min(len(data_rows) + 2, 102)):
            v = ws.cell(row=row, column=col).value
            if v:
                max_len = max(max_len, min(len(str(v)), 60))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 3

    ws.freeze_panes = "A2"
    if data_rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data_rows) + 1}"


# ── Step 1: Pull Airtable ─────────────────────────────────────────────────────

def pull_snapshot():
    print("Pulling data from Airtable...")
    all_records = []
    offset = None
    page = 1
    while True:
        params = {"pageSize": 100}
        if offset:
            params["offset"] = offset
        resp = requests.get(
            f"https://api.airtable.com/v0/{BASE_ID}/{TABLE_ID}",
            headers=HEADERS, params=params
        )
        data = resp.json()
        records = data.get("records", [])
        all_records.extend(records)
        print(f"  Page {page}: {len(records)} records")
        offset = data.get("offset")
        if not offset:
            break
        page += 1
        time.sleep(0.3)

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    snapshot = {}
    for r in all_records:
        f = r["fields"]
        email = (f.get("creator_email") or "").strip()
        domain = email.split("@")[-1].lower() if "@" in email else ""
        if domain in NON_CREATOR_DOMAINS:
            continue
        manual = f.get("action_status_manual") or ""
        if manual == "needs_no_action":
            continue
        snapshot[r["id"]] = {
            "record_id":          r["id"],
            "creator_email":      email,
            "rootlabs_inbox":     (f.get("rootlabs_email") or "").strip(),
            "thread_status":      (f.get("thread_status") or "").strip(),
            "action_status_final":(f.get("action_status_final") or "").strip(),
            "last_message_type":  (f.get("last_message_type") or "").strip(),
            "last_message_date":  (f.get("last_message_date") or "")[:19],
            "date_of_first_reply":(f.get("date_of_first_reply") or "")[:10],
            "snapshot_date":      today,
        }

    snap_path = os.path.join(SNAPSHOT_DIR, f"snapshot_{today}.json")
    with open(snap_path, "w") as fh:
        json.dump(snapshot, fh, indent=2)
    print(f"  Snapshot saved: {snap_path} ({len(snapshot)} creator records)")

    # Save immutable SOD baseline (never overwritten once created for the day)
    sod_path = os.path.join(SNAPSHOT_DIR, f"snapshot_{today}_sod.json")
    if not os.path.exists(sod_path):
        with open(sod_path, "w") as fh:
            json.dump(snapshot, fh, indent=2)
        print(f"  SOD baseline saved: {sod_path}")
    else:
        print(f"  SOD baseline already exists for today: {sod_path}")

    return snapshot, today


# ── Step 2: Load yesterday's snapshot ────────────────────────────────────────

def load_previous_snapshot(today):
    files = sorted([
        f for f in os.listdir(SNAPSHOT_DIR)
        if f.startswith("snapshot_") and f.endswith(".json") and f != f"snapshot_{today}.json"
    ], reverse=True)
    if not files:
        print("  No previous snapshot found - this is the baseline run.")
        return {}
    prev_path = os.path.join(SNAPSHOT_DIR, files[0])
    with open(prev_path) as fh:
        prev = json.load(fh)
    print(f"  Previous snapshot: {files[0]} ({len(prev)} records)")
    return prev


# ── Step 3: Diff and classify changes ─────────────────────────────────────────

def diff_snapshots(prev, curr, today):
    """
    Returns a list of change events.

    Event types:
      new_inbound        - new thread (creator replied for first time) or
                           existing thread got a new inbound message
      replied            - needs_reply  -> needs_followup_1  (done_reply sent)
      followup_1_sent    - needs_followup_1 -> needs_followup_2 (done_followup_1 sent)
      followup_2_sent    - needs_followup_2 -> needs_followup_3 (done_followup_2 sent)
      followup_3_sent    - -> abandoned                          (done_followup_3 sent)
    """
    events = []
    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    for rid, curr_rec in curr.items():
        prev_rec = prev.get(rid)
        inbox   = curr_rec["rootlabs_inbox"]
        email   = curr_rec["creator_email"]
        c_status = curr_rec["thread_status"]
        c_action = curr_rec["action_status_final"]
        c_last_date = curr_rec["last_message_date"]

        # --- New thread (didn't exist yesterday) ---
        if prev_rec is None:
            events.append({
                "detected_at":  now_str,
                "date":         today,
                "event_type":   "new_inbound",
                "creator_email": email,
                "rootlabs_inbox": inbox,
                "from_status":  "",
                "to_status":    c_status,
                "last_message_date": c_last_date,
            })
            continue

        p_status = prev_rec["thread_status"]
        p_action = prev_rec["action_status_final"]
        p_last_date = prev_rec["last_message_date"]

        # --- New inbound message on existing thread ---
        # Detected when last_message_date changed AND last_message_type is inbound
        if (c_last_date != p_last_date
                and curr_rec["last_message_type"] == "inbound"
                and p_status != "needs_reply"
                and c_status == "needs_reply"):
            events.append({
                "detected_at":  now_str,
                "date":         today,
                "event_type":   "new_inbound",
                "creator_email": email,
                "rootlabs_inbox": inbox,
                "from_status":  p_status,
                "to_status":    c_status,
                "last_message_date": c_last_date,
            })

        # --- RootLabs replied (needs_reply -> needs_followup_1) ---
        if p_status == "needs_reply" and c_action == "done_reply" and p_action != "done_reply":
            events.append({
                "detected_at":  now_str,
                "date":         today,
                "event_type":   "replied",
                "creator_email": email,
                "rootlabs_inbox": inbox,
                "from_status":  p_status,
                "to_status":    c_status,
                "last_message_date": c_last_date,
            })

        # --- Followup 1 sent (needs_followup_1 -> needs_followup_2) ---
        if p_action == "done_reply" and c_action == "done_followup_1":
            events.append({
                "detected_at":  now_str,
                "date":         today,
                "event_type":   "followup_1_sent",
                "creator_email": email,
                "rootlabs_inbox": inbox,
                "from_status":  p_status,
                "to_status":    c_status,
                "last_message_date": c_last_date,
            })

        # --- Followup 2 sent (needs_followup_2 -> needs_followup_3) ---
        if p_action == "done_followup_1" and c_action == "done_followup_2":
            events.append({
                "detected_at":  now_str,
                "date":         today,
                "event_type":   "followup_2_sent",
                "creator_email": email,
                "rootlabs_inbox": inbox,
                "from_status":  p_status,
                "to_status":    c_status,
                "last_message_date": c_last_date,
            })

        # --- Followup 3 sent (-> abandoned) ---
        if p_action == "done_followup_2" and c_action == "done_followup_3":
            events.append({
                "detected_at":  now_str,
                "date":         today,
                "event_type":   "followup_3_sent",
                "creator_email": email,
                "rootlabs_inbox": inbox,
                "from_status":  p_status,
                "to_status":    "abandoned",
                "last_message_date": c_last_date,
            })

    return events


# ── Step 4: Append to changelog CSV ──────────────────────────────────────────

CHANGELOG_FIELDS = [
    "detected_at", "date", "event_type",
    "creator_email", "rootlabs_inbox",
    "from_status", "to_status", "last_message_date"
]

def append_changelog(events):
    file_exists = os.path.exists(CHANGELOG_PATH)
    with open(CHANGELOG_PATH, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CHANGELOG_FIELDS)
        if not file_exists:
            writer.writeheader()
        for e in events:
            writer.writerow(e)
    print(f"  Changelog: {len(events)} new events appended -> {CHANGELOG_PATH}")


# ── Step 5: Load full changelog ───────────────────────────────────────────────

def load_changelog():
    if not os.path.exists(CHANGELOG_PATH):
        return []
    with open(CHANGELOG_PATH, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


# ── Step 6: Build daily summary Excel ────────────────────────────────────────

def build_summary_excel(current_snapshot, changelog_rows):
    print("  Building daily_summary.xlsx...")
    wb = Workbook()

    # ── Sheet 1: Current Queue Snapshot ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Current Queue"

    headers1 = [
        "Inbox",
        "Needs Reply", "Needs Followup 1", "Needs Followup 2", "Needs Followup 3",
        "Total Active"
    ]
    data1 = []
    totals = defaultdict(int)

    for inbox in INBOXES:
        recs = [v for v in current_snapshot.values() if v["rootlabs_inbox"] == inbox]
        nr   = sum(1 for r in recs if r["thread_status"] == "needs_reply")
        nf1  = sum(1 for r in recs if r["thread_status"] == "needs_followup_1")
        nf2  = sum(1 for r in recs if r["thread_status"] == "needs_followup_2")
        nf3  = sum(1 for r in recs if r["thread_status"] == "needs_followup_3")
        total = nr + nf1 + nf2 + nf3
        data1.append([inbox, nr, nf1, nf2, nf3, total])
        for key, val in zip(["nr","nf1","nf2","nf3","total"], [nr,nf1,nf2,nf3,total]):
            totals[key] += val

    # Totals row
    data1.append([
        "TOTAL",
        totals["nr"], totals["nf1"], totals["nf2"], totals["nf3"], totals["total"]
    ])

    row_fills1 = [None] * (len(data1) - 1) + [TOTAL_FILL]
    style_sheet(ws1, headers1, data1, row_fills1)

    # Bold the totals row
    for col in range(1, len(headers1) + 1):
        ws1.cell(row=len(data1) + 1, column=col).font = BOLD_FONT

    # ── Sheet 2: Daily Execution Log (from changelog) ─────────────────────────
    ws2 = wb.create_sheet("Daily Execution Log")

    # Group changelog by date
    by_date = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    all_dates = sorted(set(r["date"] for r in changelog_rows), reverse=True)

    for row in changelog_rows:
        date   = row["date"]
        inbox  = row["rootlabs_inbox"]
        etype  = row["event_type"]
        by_date[date][inbox][etype] += 1

    headers2 = [
        "Date", "Inbox",
        "New Inbounds",
        "Replies Sent (needs_reply actioned)",
        "Followup 1 Sent",
        "Followup 2 Sent",
        "Followup 3 Sent",
        "Total Actions"
    ]

    data2 = []
    for date in all_dates:
        for inbox in INBOXES:
            counts = by_date[date].get(inbox, {})
            new_in = counts.get("new_inbound", 0)
            replied = counts.get("replied", 0)
            fu1     = counts.get("followup_1_sent", 0)
            fu2     = counts.get("followup_2_sent", 0)
            fu3     = counts.get("followup_3_sent", 0)
            total_actions = replied + fu1 + fu2 + fu3
            if new_in + total_actions > 0:
                data2.append([date, inbox, new_in, replied, fu1, fu2, fu3, total_actions])

        # Date subtotal
        all_inboxes_counts = by_date[date]
        new_in_t  = sum(v.get("new_inbound",0) for v in all_inboxes_counts.values())
        replied_t = sum(v.get("replied",0) for v in all_inboxes_counts.values())
        fu1_t     = sum(v.get("followup_1_sent",0) for v in all_inboxes_counts.values())
        fu2_t     = sum(v.get("followup_2_sent",0) for v in all_inboxes_counts.values())
        fu3_t     = sum(v.get("followup_3_sent",0) for v in all_inboxes_counts.values())
        total_t   = replied_t + fu1_t + fu2_t + fu3_t
        if new_in_t + total_t > 0:
            data2.append([date, "ALL INBOXES", new_in_t, replied_t, fu1_t, fu2_t, fu3_t, total_t])

    if not data2:
        data2 = [["No data yet - run again tomorrow to see changes", "", "", "", "", "", "", ""]]

    style_sheet(ws2, headers2, data2)

    # Highlight "ALL INBOXES" subtotal rows
    for r_idx in range(2, len(data2) + 2):
        if ws2.cell(row=r_idx, column=2).value == "ALL INBOXES":
            for c in range(1, len(headers2) + 1):
                ws2.cell(row=r_idx, column=c).fill = TOTAL_FILL
                ws2.cell(row=r_idx, column=c).font = BOLD_FONT

    # ── Sheet 3: All Changelog Events ────────────────────────────────────────
    ws3 = wb.create_sheet("Full Changelog")

    headers3 = ["Date", "Detected At", "Event Type", "Creator Email",
                 "RootLabs Inbox", "From Status", "To Status", "Last Message Date"]
    data3 = []

    event_label = {
        "new_inbound":     "New Inbound",
        "replied":         "Reply Sent",
        "followup_1_sent": "Followup 1 Sent",
        "followup_2_sent": "Followup 2 Sent",
        "followup_3_sent": "Followup 3 Sent",
    }
    event_fill = {
        "new_inbound":     AMBER_FILL,
        "replied":         GREEN_FILL,
        "followup_1_sent": None,
        "followup_2_sent": None,
        "followup_3_sent": RED_FILL,
    }

    row_fills3 = []
    for row in sorted(changelog_rows, key=lambda x: x["detected_at"], reverse=True):
        etype = row["event_type"]
        data3.append([
            row["date"],
            row["detected_at"],
            event_label.get(etype, etype),
            row["creator_email"],
            row["rootlabs_inbox"],
            row["from_status"],
            row["to_status"],
            row["last_message_date"],
        ])
        row_fills3.append(event_fill.get(etype))

    if not data3:
        data3 = [["No changelog events yet", "", "", "", "", "", "", ""]]
        row_fills3 = [None]

    style_sheet(ws3, headers3, data3, row_fills3)

    wb.save(SUMMARY_PATH)
    print(f"  Saved: {SUMMARY_PATH}")


# ── Scorecard ─────────────────────────────────────────────────────────────────

def compute_scorecard_metrics(date_str, sod_snap, curr_snap):
    """Compute full daily scorecard metrics by diffing SOD vs current snapshot."""
    new_inbound = 0
    replied = followup_1_sent = followup_2_sent = followup_3_sent = 0

    sod_ids  = set(sod_snap.keys())
    curr_ids = set(curr_snap.keys())

    for rid in curr_ids - sod_ids:
        if curr_snap[rid].get("last_message_type") == "inbound":
            new_inbound += 1

    for rid, sod_rec in sod_snap.items():
        if rid not in curr_snap:
            continue
        curr_rec = curr_snap[rid]
        p_status = sod_rec["thread_status"]
        c_status = curr_rec["thread_status"]
        p_action = sod_rec["action_status_final"]
        c_action = curr_rec["action_status_final"]
        p_date   = sod_rec["last_message_date"]
        c_date   = curr_rec["last_message_date"]

        if (p_status != "needs_reply" and c_status == "needs_reply"
                and c_date != p_date and curr_rec.get("last_message_type") == "inbound"):
            new_inbound += 1
        if p_status == "needs_reply" and c_action == "done_reply" and p_action != "done_reply":
            replied += 1
        if p_action == "done_reply" and c_action == "done_followup_1":
            followup_1_sent += 1
        if p_action == "done_followup_1" and c_action == "done_followup_2":
            followup_2_sent += 1
        if p_action == "done_followup_2" and c_action == "done_followup_3":
            followup_3_sent += 1

    sod_nr  = sum(1 for r in sod_snap.values() if r["thread_status"] == "needs_reply")
    sod_nf1 = sum(1 for r in sod_snap.values() if r["thread_status"] == "needs_followup_1")
    sod_nf2 = sum(1 for r in sod_snap.values() if r["thread_status"] == "needs_followup_2")
    sod_nf3 = sum(1 for r in sod_snap.values() if r["thread_status"] == "needs_followup_3")
    sod_total = sod_nr + sod_nf1 + sod_nf2 + sod_nf3
    total_actioned = replied + followup_1_sent + followup_2_sent + followup_3_sent

    def rate(n, d):
        return round(n / d * 100, 1) if d > 0 else 0.0

    return {
        "date":                  date_str,
        "saved_at":              datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        "sod_needs_reply":       sod_nr,
        "sod_needs_followup_1":  sod_nf1,
        "sod_needs_followup_2":  sod_nf2,
        "sod_needs_followup_3":  sod_nf3,
        "sod_total":             sod_total,
        "new_inbounds":          new_inbound,
        "replied":               replied,
        "followup_1_sent":       followup_1_sent,
        "followup_2_sent":       followup_2_sent,
        "followup_3_sent":       followup_3_sent,
        "total_actioned":        total_actioned,
        "reply_rate_pct":        rate(replied, sod_nr),
        "fu1_rate_pct":          rate(followup_1_sent, sod_nf1),
        "fu2_rate_pct":          rate(followup_2_sent, sod_nf2),
        "fu3_rate_pct":          rate(followup_3_sent, sod_nf3),
        "overall_rate_pct":      rate(total_actioned, sod_total),
    }


def save_scorecard_row(metrics):
    """Append or overwrite today's row in daily_scorecards.csv."""
    date_str = metrics["date"]
    rows = []
    if os.path.exists(SCORECARD_PATH):
        with open(SCORECARD_PATH, newline="", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
    rows = [r for r in rows if r.get("date") != date_str]
    rows.append(metrics)
    rows.sort(key=lambda r: r["date"])
    with open(SCORECARD_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=SCORECARD_FIELDS)
        writer.writeheader()
        writer.writerows(rows)
    print(f"  Scorecard row saved for {date_str} -> {SCORECARD_PATH}")


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    eod_mode = "--eod" in sys.argv

    print("=" * 60)
    if eod_mode:
        print("Daily Outreach Tracker — EOD Scorecard Mode")
    else:
        print("Daily Outreach Tracker — Morning Run")
    print(f"Run time: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")
    print("=" * 60)

    if eod_mode:
        # ── EOD: save scorecard only, do not touch changelog ─────────────────
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        sod_path = os.path.join(SNAPSHOT_DIR, f"snapshot_{today}_sod.json")
        if not os.path.exists(sod_path):
            print("ERROR: No SOD snapshot found for today. Run the morning tracker first.")
            sys.exit(1)
        with open(sod_path) as fh:
            sod_snap = json.load(fh)
        print(f"  SOD baseline loaded: {len(sod_snap)} records")

        print("\nPulling current Airtable state for EOD diff...")
        curr_snapshot, _ = pull_snapshot()

        print("\nComputing EOD scorecard...")
        metrics = compute_scorecard_metrics(today, sod_snap, curr_snapshot)
        save_scorecard_row(metrics)

        print(f"\n{'─' * 40}")
        print(f"  EOD Scorecard — {today}")
        print(f"{'─' * 40}")
        print(f"  SOD total queue:   {metrics['sod_total']}")
        print(f"  New inbounds:      {metrics['new_inbounds']}")
        print(f"  Replies sent:      {metrics['replied']}  ({metrics['reply_rate_pct']}% of needs_reply)")
        print(f"  Followup 1 sent:   {metrics['followup_1_sent']}  ({metrics['fu1_rate_pct']}% of FU1 queue)")
        print(f"  Followup 2 sent:   {metrics['followup_2_sent']}  ({metrics['fu2_rate_pct']}% of FU2 queue)")
        print(f"  Total actioned:    {metrics['total_actioned']}  ({metrics['overall_rate_pct']}% overall)")
        print(f"{'─' * 40}")

    else:
        # ── Morning run: snapshot + changelog + Excel ─────────────────────────
        # 1. Pull fresh snapshot
        curr_snapshot, today = pull_snapshot()

        # 2. Load previous snapshot
        print("\nLoading previous snapshot...")
        prev_snapshot = load_previous_snapshot(today)

        # 3. Diff
        print("\nDetecting changes...")
        events = diff_snapshots(prev_snapshot, curr_snapshot, today)
        print(f"  {len(events)} change events detected")
        if events:
            for etype, count in Counter(e["event_type"] for e in events).most_common():
                print(f"    {etype}: {count}")

        # 4. Append changelog
        print("\nUpdating changelog...")
        append_changelog(events)

        # 5. Load full changelog
        changelog = load_changelog()
        print(f"  Total changelog entries so far: {len(changelog)}")

        # 6. Build Excel
        print("\nBuilding summary Excel...")
        build_summary_excel(curr_snapshot, changelog)

        print("\n" + "=" * 60)
        print("Done.")
        print(f"  Queue snapshot:  snapshots/snapshot_{today}.json")
        print(f"  Changelog:       changelog.csv")
        print(f"  Summary Excel:   daily_summary.xlsx")
        print("=" * 60)
